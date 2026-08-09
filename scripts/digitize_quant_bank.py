#!/usr/bin/env python3
"""Build the Quant/LR mock bank from the supplied schedule and source PDFs.

The source PDFs contain:
  * the exercise questions and answer choices,
  * shared Directions/Data blocks used by question ranges, and
  * answer keys on the final page.

This script keeps those three layers connected. It deliberately fails when a
scheduled source question cannot be resolved instead of filling gaps with
placeholder content.
"""

from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
QUANT_DIR = ROOT / "Quant"
SCHEDULE_PATH = QUANT_DIR / "Quant_and_LR_31Day_Mock_Drill_Schedule.xlsx"
OUTPUT_PATH = ROOT / "src" / "data" / "question_bank.json"
PUBLIC_PDF_DIR = ROOT / "public" / "quant-sources"


@dataclass(frozen=True)
class Section:
    name: str
    marker: str
    count: int


MANIFEST: dict[str, list[Section]] = {
    "SM1001905_Chapter-10_Deductions.pdf": [
        Section("Exercise 10(a)", r"Exercise\s*[–-]\s*10\(a\)", 30),
        Section("Exercise 10(b)", r"Exercise\s*[–-]\s*10\(b\)", 30),
    ],
    "SM1001905_Chapter-12_Allareas.pdf": [
        Section("Exercise 12(a)", r"Exercise\s*[–-]\s*12\(a\)", 30),
        Section("Exercise 12(b)", r"Exercise\s*[–-]\s*12\(b\)", 30),
        Section("Exercise 12(c)", r"Exercise\s*[–-]\s*12\(c\)", 30),
    ],
    "SM1001905_Chapter-1_LinearArrangement.pdf": [
        Section("Exercise 1(a)", r"Exercise\s*[–-]\s*1\(a\)", 30),
        Section("Exercise 1(b)", r"Exercise\s*[–-]\s*1\(b\)", 30),
    ],
    "SM1001905_Chapter-5_OrderingandSequence.pdf": [
        Section("Exercise 5(a)", r"Exercise\s*[–-]\s*5\(a\)", 30),
        Section("Exercise 5(b)", r"Exercise\s*[–-]\s*5\(b\)", 30),
    ],
    "SM1001906_Chapter-2_RPV.pdf": [
        Section("Concept Review", r"Concept Review Questions", 30),
        Section("Exercise 2(a)", r"Exercise\s*[–-]\s*2\(a\)", 30),
        Section("Exercise 2(b)", r"Exercise\s*[–-]\s*2\(b\)", 45),
    ],
    "SM1001906_Chapter-7_(AMA).pdf": [
        Section("Concept Review", r"Concept Review Questions", 35),
        Section("Exercise 7(a)", r"Exercise\s*[–-]\s*7\(a\)", 30),
        Section("Exercise 7(b)", r"Exercise\s*[–-]\s*7\(b\)", 45),
    ],
    "SM1001907_Chapter-1(NumberAndLetterSeries).pdf": [
        Section("Exercise 1(a)", r"Exercise\s*[–-]\s*1\(a\)", 75),
        Section("Exercise 1(b)", r"Exercise\s*[–-]\s*1\(b\)", 50),
    ],
    "SM1001907_Chapter-10(AnalyticalPuzzles).pdf": [
        Section("Exercise 10(a)", r"Exercise\s*[–-]\s*10\(a\)", 50),
        Section("Exercise 10(b)", r"Exercise\s*[–-]\s*10\(b\)", 50),
        Section("Exercise 10(c)", r"Exercise\s*[–-]\s*10\(c\)", 50),
        Section("Exercise 10(d)", r"Exercise\s*[–-]\s*10\(d\)", 70),
    ],
    "SM1001907_Chapter-2(Analogies).pdf": [
        Section("Exercise 2(a)", r"Exercise\s*[–-]\s*2\(a\)", 50),
        Section("Exercise 2(b)", r"Exercise\s*[–-]\s*2\(b\)", 50),
    ],
    "SM1001907_Chapter-3(CodingDecoding).pdf": [
        Section("Exercise 3(a)", r"Exercise\s*[–-]\s*3\(a\)", 50),
        Section("Exercise 3(b)", r"Exercise\s*[–-]\s*3\(b\)", 40),
    ],
    "SM1001907_Chapter-4(Oddmanout).pdf": [
        Section("Exercise 4", r"Exercise\s*[–-]\s*4\b", 50),
    ],
    "SM1001907_Chapter-7(SymbolsNotations).pdf": [
        Section("Exercise 7", r"Exercise\s*[–-]\s*7\b", 80),
    ],
    "SM1001907_Chapter-8(DirectionSense).pdf": [
        Section("Exercise 8", r"Exercise\s*[–-]\s*8\b", 50),
    ],
    "SM1001907_Chapter-9(BloodRelations).pdf": [
        Section("Exercise 9", r"Exercise\s*[–-]\s*9\b", 50),
    ],
}


CATEGORY_BY_TOPIC = {
    "Deductions": "Logical Reasoning",
    "All Areas (Puzzles/Arrangements)": "Analytical Puzzles",
    "Linear Arrangement": "Logical Reasoning",
    "Ordering & Sequence": "Logical Reasoning",
    "Ratio, Proportion & Variation": "Quantitative Techniques",
    "Averages, Mixtures & Alligations": "Quantitative Techniques",
    "Number & Letter Series": "Logical Reasoning",
    "Analytical Puzzles": "Analytical Puzzles",
    "Analogies": "Logical Reasoning",
    "Coding & Decoding": "Logical Reasoning",
    "Odd Man Out": "Logical Reasoning",
    "Symbols & Notations": "Logical Reasoning",
    "Direction Sense": "Logical Reasoning",
    "Blood Relations": "Logical Reasoning",
}


GENERIC_DIRECTIONS = (
    "select the correct alternative",
    "select the correct choice",
    "find the odd man out",
    "complete the following series",
    "for the multiple choice questions",
    "in each question below",
    "in each of the following questions",
    "each question below has",
    "each of these questions consists",
    "each of the following questions consists",
    "choose from the five diagrams",
)


VISUAL_CUES = (
    "following diagram",
    "diagram given",
    "given figure",
    "given figures",
    "bar graph",
    "pie chart",
    "line graph",
    "venn diagram",
    "shown below",
)

# The printed key marks this numeric-response question as "B", while the page
# clearly shows a numeric answer box. Solving the printed ratios gives 20.
ANSWER_OVERRIDES = {
    (
        "SM1001906_Chapter-2_RPV.pdf",
        "Exercise 2(b)",
        14,
    ): "20",
}


PURE_QUANT_TOPICS = {
    "Ratio, Proportion & Variation",
    "Averages, Mixtures & Alligations",
}


# pypdf extracts ordinary prose from these source files reliably, but vertically
# laid-out maths (fractions, exponents, radicals and mixed numbers) can be emitted
# in the wrong reading order. Keep the visually verified corrections beside the
# digitizer so rebuilding the bank cannot reintroduce the damaged expressions.
CONTENT_OVERRIDES: dict[tuple[str, str, int], dict[str, Any]] = {
    ("SM1001906_Chapter-2_RPV.pdf", "Concept Review", 1): {
        "questionText": "If a = (9/8)b, then (3b)/(4a) = ?",
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Concept Review", 2): {
        "questionText": "If 2.4p = 0.08q, then (q + p)/(q - p) = ?",
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Concept Review", 3): {
        "questionText": "If a : b = 2 : 3, find (3a + 4b)/(4a + 5b).",
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Concept Review", 22): {
        "questionText": (
            "Three positive numbers p, q and r satisfy "
            "(q + r)/p = (p + r)/q = (p + q)/r = K. Find K."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(a)", 1): {
        "questionText": (
            "If a : b = 3 : 7, what is the value of (4a + 5b)/(2a + 2b)?"
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(a)", 6): {
        "options": [
            "(p + q)/3",
            "(3p - q)/4",
            "(p - q)/(p + q)",
            "(q - 3p)/4",
        ],
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(a)", 16): {
        "questionText": (
            "If k = (a + c)/(b + d) = (c + e)/(d + f) = "
            "(a + e)/(b + f), when all quantities are positive, then which "
            "of the following must be true?"
        ),
        "options": ["k = e/f", "k = a/b", "k = c/d", "All of the above"],
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(a)", 18): {
        "questionText": (
            "The mean proportional between two numbers is 9 and the third "
            "proportional of the two numbers is 243. Find the larger of the two numbers."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(a)", 19): {
        "questionText": (
            "P and Q are distinct two-digit numbers. Pₛ and Qₛ denote the sums "
            "of the digits in P and Q respectively. If P/Pₛ = Q/Qₛ, find the "
            "minimum possible value of Pₛ + Qₛ."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 2): {
        "questionText": (
            "Which of the following represents a possible value of p : q satisfying "
            "(20p² - 40pq)/(pq + 4q²) = 20?"
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 3): {
        "questionText": (
            "If a : b = 3 : 4 and c : d = 2 : 3, find "
            "(a³c² + b³d²)/(ab²d² + a²bcd)."
        ),
        "options": ["19/3", "19/4", "19/9", "19/18"],
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 11): {
        "options": ["1/3", "1/5", "1/4", "2/5"],
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 14): {
        "questionText": (
            "The ratio of the prices of tea last year and this year is 5 : 6. "
            "The ratio of the prices of coffee last year and this year is 7 : 8. "
            "The sum of the prices of a kg of tea and a kg of coffee this year is "
            "₹48. Find the price of tea (in ₹) last year if it was 20/21 of the "
            "price of coffee last year."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 15): {
        "questionText": (
            "₹117 was supposed to be divided among Rohan, Sohan and Mohan in the "
            "ratio 2 : 3 : 4. By mistake, it was divided in the ratio "
            "1/2 : 1/3 : 1/4. Find Rohan's loss or gain due to this mistake (in ₹)."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 24): {
        "questionText": (
            "If (a² + c²)/(a + c) = (b² + c²)/(b + c) = k and a ≠ b, "
            "which of the following is equal to k?"
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 25): {
        "questionText": (
            "If (p + q)/r = (q + r)/p = (p + r)/q = k, find the sum of all "
            "possible values of k."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 26): {
        "questionText": (
            "If (2x² - 4x + 3)/(4x - 3) = "
            "(2x² - 3x + 5)/(3x - 5), find the value(s) of x."
        ),
    },
    ("SM1001906_Chapter-2_RPV.pdf", "Exercise 2(b)", 30): {
        "options": ["2 : 1", "4 : 1", "3 : 1", "√2 : 1"],
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Concept Review", 25): {
        "questionText": (
            "A vessel has 20 litres of a mixture of milk and water having 60% milk. "
            "Five litres of pure milk is added to the vessel. Find the percentage "
            "of milk in the new solution."
        ),
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(a)", 6): {
        "options": [
            "1/n = (x - y)/(a - b)",
            "1/n = (a - b)/(y - x)",
            "1/n = (x + y)/(a + b)",
            "1/n = (a + b)/(x + y)",
        ],
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(a)", 11): {
        "questionText": (
            "Four classes take an algebra test: Class A has 32 students with an "
            "average of 83; Class B has 58 students with an average of 76; Class C "
            "has 82 students with an average of 85; and Class D has 48 students with "
            "an average of 90. Find the combined average score of all four classes."
        ),
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(a)", 17): {
        "options": [
            "₹128/7 per kg",
            "₹120/7 per kg",
            "₹141/7 per kg",
            "₹149/7 per kg",
        ],
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(a)", 21): {
        "questionText": (
            "Raju bought 40 kg of coffee powder costing ₹240/kg and 60 kg of another "
            "variety costing ₹360/kg. He mixed them and sold 4/5 of the mixture at "
            "₹360/kg and the rest at ₹270/kg. What overall profit percentage did he make?"
        ),
        "options": ["9 8/13%", "11 4/7%", "5 5/11%", "17 2/19%"],
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(a)", 29): {
        "questionText": (
            "A vessel contains 100 litres of a milk-and-water mixture with 90% milk. "
            "Each time, 10% of the contents is withdrawn and replaced with an equal "
            "amount of water. What is the minimum number of repetitions needed for "
            "the milk concentration to fall below 66 2/3%?"
        ),
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(b)", 4): {
        "questionText": (
            "A basketball player played nine matches and averaged 16 points per match. "
            "His score in the i-th match was two less than in the (i - 1)-th match. "
            "Find the average points scored in the second and eighth matches."
        ),
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(b)", 9): {
        "options": ["9", "10", "11 7/11", "11 5/11"],
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(b)", 36): {
        "questionText": (
            "Three varieties of rice A, B and C costing ₹6/kg, ₹9/kg and ₹12/kg are "
            "mixed in a certain ratio. The mixture is sold at a profit of 66 2/3% for "
            "₹15/kg. Of the 100 kg mixture, 50 kg is variety B. Find the quantity of "
            "variety A (in kg)."
        ),
    },
    ("SM1001906_Chapter-7_(AMA).pdf", "Exercise 7(b)", 42): {
        "questionText": (
            "A certain alloy contains lead, copper and tin. How many kilograms of tin "
            "are contained in 60 kilograms of the alloy? I. By weight, the alloy is "
            "2/5 lead and 3/16 copper. II. By volume, the alloy is 1/3 lead and 1/3 copper."
        ),
    },
}


QUANT_TEXT_REPAIRS = {
    "corre ct": "correct",
    "corr ect": "correct",
    "wri te": "write",
    "adde d": "added",
    "i s": "is",
    "t he": "the",
    "followi ng": "following",
    "so n": "son",
    "o f": "of",
    "classe s": "classes",
    "Fi nd": "Find",
    "salar y": "salary",
    "proportion al": "proportional",
    "cov er": "cover",
    "wat er": "water",
    "pri ce": "price",
    "g ets": "gets",
    "thr ee": "three",
    "va riety": "variety",
    "T he": "The",
    "rati o": "ratio",
    "travell ed": "travelled",
    "t ime": "time",
    "sufficien t": "sufficient",
    "als o": "also",
    "fix ed": "fixed",
    "fr om": "from",
    "quant ity": "quantity",
    "numbe r": "number",
    "bel ow": "below",
    "k g": "kg",
    "th rice": "thrice",
    "son s": "sons",
    "an d": "and",
    "tot al": "total",
    "ye ar": "year",
    "te a": "tea",
    "gi rls": "girls",
    "scholar s": "scholars",
    "scho ol": "school",
    "wha t": "what",
    "su m": "sum",
    "are as": "areas",
    "questio ns": "questions",
    "thei r": "their",
    "le ngths": "lengths",
    "squa re": "square",
    "va lue": "value",
    "mathemat ics": "mathematics",
    "Mathema tics": "Mathematics",
    "Engli sh": "English",
    "lit res": "litres",
    "stude nt": "student",
    "resu lt": "result",
    "mi lk": "milk",
    "solutio n": "solution",
    "conte nts": "contents",
    "currentl y": "currently",
    "variet y": "variety",
    "ad ded": "added",
    "w ith": "with",
    "averag e": "average",
    "inter est": "interest",
    "p. a.": "p.a.",
    "opthe": "of the",
    "th an": "than",
    "ag o": "ago",
    "te sts": "tests",
    "avera ge": "average",
    "th e": "the",
    "cost pric e": "cost price",
    "to tal": "total",
    "re sult": "result",
    "var iation": "variation",
    "par tly": "partly",
    "staf f": "staff",
    "uni ts": "units",
    "sta ge": "stage",
    "weig ht": "weight",
    "aft er": "after",
    "anot her": "another",
    "dire ctly": "directly",
    "equ al": "equal",
    "i t": "it",
    "ra tio": "ratio",
    "rat io": "ratio",
    "a re": "are",
    "scor e": "score",
    "shoul d": "should",
    "sufficie nt": "sufficient",
    "teach er": "teacher",
    "thi rd": "third",
    "tota l": "total",
    "wate r": "water",
    "weigh t": "weight",
}


def normalize_quant_content(value: Any) -> Any:
    if isinstance(value, str):
        value = value.replace("`", "₹").replace("ΙΙ", "II").replace("Ι", "I")
        for damaged, repaired in QUANT_TEXT_REPAIRS.items():
            value = re.sub(
                rf"(?<![A-Za-z]){re.escape(damaged)}(?![A-Za-z])",
                repaired,
                value,
            )
        value = re.sub(r"\s+([,.;:?!])", r"\1", value)
        return re.sub(r"\s+", " ", value).strip()
    if isinstance(value, list):
        return [normalize_quant_content(item) for item in value]
    if isinstance(value, dict):
        return {key: normalize_quant_content(item) for key, item in value.items()}
    return value


def clean_lines(text: str) -> str:
    kept: list[str] = []
    for line in text.replace("\u00a0", " ").splitlines():
        if "Triumphant Institute of Management Education" in line:
            continue
        if "Tel :" in line and "time4education" in line:
            continue
        if re.fullmatch(r"\s*SM\d+/\d+\s*", line):
            continue
        kept.append(line.rstrip())
    return "\n".join(kept)


def compact(text: str) -> str:
    text = re.sub(r"\[\[PAGE\s+\d+\]\]", " ", text)
    text = text.replace("\uf0e4", "→").replace("\uf0a3", "£")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+([,.;:?!])", r"\1", text)
    text = re.sub(r"([(\[])\s+", r"\1", text)
    text = re.sub(r"\s+([)\]])", r"\1", text)
    return text.strip()


def slug(value: str) -> str:
    value = re.sub(r"\.pdf$", "", value, flags=re.I)
    value = re.sub(r"[^a-z0-9]+", "-", value.lower())
    return value.strip("-")


def load_pdf_text(pdf_path: Path) -> tuple[str, int]:
    reader = PdfReader(str(pdf_path))
    parts: list[str] = []
    for page_number, page in enumerate(reader.pages, 1):
        parts.append(
            f"\n[[PAGE {page_number}]]\n{clean_lines(page.extract_text() or '')}"
        )
    return "".join(parts), len(reader.pages)


def find_key_position(text: str) -> int:
    matches = list(re.finditer(r"(?m)^\s*Key\s*$", text, re.I))
    if not matches:
        raise ValueError("Answer-key marker not found")
    return matches[-1].start()


def locate_sections(
    full_text: str, sections: list[Section], key_position: int
) -> list[tuple[Section, int, int]]:
    starts: list[int] = []
    for section in sections:
        candidates = [
            match
            for match in re.finditer(section.marker, full_text, re.I)
            if match.start() < key_position
        ]
        if not candidates:
            raise ValueError(f"Section marker not found: {section.name}")
        starts.append(candidates[-1].start())

    located: list[tuple[Section, int, int]] = []
    for index, section in enumerate(sections):
        end = starts[index + 1] if index + 1 < len(starts) else key_position
        located.append((section, starts[index], end))
    return located


def question_marker(section_text: str, question_number: int, start: int) -> re.Match[str]:
    # A single source typo prints "74 Which..." rather than "74. Which...".
    pattern = re.compile(
        rf"(?m)^\s*{question_number}(?:\.\s*|\s+(?=[A-Z]))"
    )
    match = pattern.search(section_text, start)
    if not match:
        raise ValueError(f"Question {question_number} marker not found")
    return match


def page_before(text: str, position: int, default_page: int = 1) -> int:
    matches = list(re.finditer(r"\[\[PAGE\s+(\d+)\]\]", text[:position]))
    return int(matches[-1].group(1)) if matches else default_page


def split_options(raw_chunk: str) -> tuple[str, list[str]]:
    markers = list(re.finditer(r"\(([A-E])\)", raw_chunk))
    marker_letters = [marker.group(1) for marker in markers]

    # Preserve visual order for source typos such as A, C, C, D, E.
    if 2 <= len(markers) <= 5 and marker_letters != list("ABCDE")[: len(markers)]:
        question_text = compact(raw_chunk[: markers[0].start()])
        options = []
        for index, marker in enumerate(markers):
            end = markers[index + 1].start() if index + 1 < len(markers) else len(raw_chunk)
            options.append(compact(raw_chunk[marker.end() : end]))
        return question_text, options

    selected: list[re.Match[str]] = []
    expected = ord("A")
    for marker in markers:
        letter = marker.group(1)
        if ord(letter) == expected:
            selected.append(marker)
            expected += 1
            if expected > ord("E"):
                break

    if not selected:
        return compact(raw_chunk), []

    question_text = compact(raw_chunk[: selected[0].start()])
    options: list[str] = []
    for index, marker in enumerate(selected):
        end = selected[index + 1].start() if index + 1 < len(selected) else len(raw_chunk)
        options.append(compact(raw_chunk[marker.end() : end]))
    return question_text, options


def split_numbered_options(raw_chunk: str) -> tuple[str, list[str]]:
    markers = list(re.finditer(r"\(([1-5])\)", raw_chunk))
    if len(markers) < 2:
        return compact(raw_chunk), []
    question_text = compact(raw_chunk[: markers[0].start()])
    options: list[str] = []
    for index, marker in enumerate(markers[:5]):
        end = markers[index + 1].start() if index + 1 < len(markers) else len(raw_chunk)
        options.append(compact(raw_chunk[marker.end() : end]))
    return question_text, options


def split_multi_part_question(
    raw_chunk: str, answer_raw: str
) -> tuple[str, list[dict[str, Any]]]:
    answer_match = re.fullmatch(
        r"\(i\)\s*(.+?)\s*\(ii\)\s*(.+)", answer_raw, re.I
    )
    if not answer_match:
        raise ValueError(f"Unrecognized multi-part answer: {answer_raw!r}")

    part_markers = list(
        re.finditer(r"\((?:i|Ι)\)|\((?:ii|ΙΙ)\)", raw_chunk, re.I)
    )
    if len(part_markers) < 2:
        raise ValueError("Multi-part answer found without two source sub-questions")

    main_text = compact(raw_chunk[: part_markers[0].start()])
    answers = [compact(answer_match.group(1)), compact(answer_match.group(2))]
    sub_questions: list[dict[str, Any]] = []
    for index, marker in enumerate(part_markers[:2]):
        end = (
            part_markers[index + 1].start()
            if index + 1 < len(part_markers)
            else len(raw_chunk)
        )
        part_text, part_options = split_options(raw_chunk[marker.end() : end])
        part_answer = answers[index]
        is_choice = bool(re.fullmatch(r"[A-E]", part_answer))
        sub_questions.append(
            {
                "label": "(i)" if index == 0 else "(ii)",
                "questionText": part_text,
                "options": part_options,
                "questionType": "MCQ" if is_choice else "NUMERIC",
                "correctOption": part_answer if is_choice else None,
                "numericAnswer": None if is_choice else part_answer,
            }
        )
    return main_text, sub_questions


def parse_answer_section(
    key_text: str,
    section: Section,
    section_index: int,
    all_sections: list[Section],
) -> dict[int, str]:
    markers: list[tuple[Section, int]] = []
    for item in all_sections:
        candidates = list(re.finditer(item.marker, key_text, re.I))
        if candidates:
            markers.append((item, candidates[-1].start()))

    if len(all_sections) == 1 and not markers:
        body = key_text
    else:
        current = next((position for item, position in markers if item.name == section.name), None)
        if current is None:
            raise ValueError(f"Answer-key section not found: {section.name}")
        later = sorted(position for _, position in markers if position > current)
        body = key_text[current : later[0] if later else len(key_text)]

    normalized = compact(body)
    hits: list[tuple[int, int, int]] = []
    cursor = 0
    for number in range(1, section.count + 1):
        # A handful of printed keys omit the full stop after the number.
        hit = re.search(
            rf"(?<!\d){number}(?:\.\s*|\s+(?=[A-E(]))",
            normalized[cursor:],
        )
        if not hit:
            continue
        absolute_start = cursor + hit.start()
        absolute_end = cursor + hit.end()
        hits.append((number, absolute_start, absolute_end))
        cursor = absolute_end

    answers: dict[int, str] = {}
    for index, (number, _, start) in enumerate(hits):
        end = (
            hits[index + 1][1]
            if index + 1 < len(hits)
            else len(normalized)
        )
        answer = normalized[start:end].strip()
        # The next exercise heading can remain at the end of the last answer.
        answer = re.split(r"\bExercise\s*[–-]", answer, maxsplit=1)[0].strip()
        if answer:
            answers[number] = answer
    return answers


def is_substantive_context(text: str) -> bool:
    value = compact(text)
    if not value:
        return False
    lowered = value.lower()
    if any(lowered.startswith(prefix) for prefix in GENERIC_DIRECTIONS):
        # Long generic blocks can still define answer rules or transformations.
        return any(
            token in lowered
            for token in (
                "means",
                "code",
                "following rules",
                "following directives",
                "based on",
            )
        )
    return len(value) >= 45


def parse_direction_ranges(
    section_text: str,
    question_starts: dict[int, int],
    pdf_name: str,
    section: Section,
    default_page: int,
) -> dict[int, dict[str, Any]]:
    assignments: dict[int, dict[str, Any]] = {}
    pattern = re.compile(
        r"Directions?\s+for\s+questions?\s+"
        r"(?:for\s+)?(\d+)"
        r"(?:\s*(?:to|and|[-–—])\s*(\d+))?\s*:?",
        re.I,
    )

    for match in pattern.finditer(section_text):
        start_number = int(match.group(1))
        end_number = int(match.group(2) or match.group(1))
        if start_number not in question_starts:
            continue
        first_question_position = question_starts[start_number]
        if match.start() > first_question_position:
            continue
        body = compact(section_text[match.end() : first_question_position])
        substantive = is_substantive_context(body)
        stimulus_id = (
            f"{slug(pdf_name)}-{slug(section.name)}-q{start_number}-q{end_number}"
        )
        source_page = page_before(section_text, match.start(), default_page)
        visual = any(cue in body.lower() for cue in VISUAL_CUES)
        for question_number in range(start_number, end_number + 1):
            if question_number not in question_starts:
                continue
            assignments[question_number] = {
                "stimulusId": stimulus_id,
                "stimulusType": (
                    "visual" if visual else ("passage" if substantive else "directions")
                ),
                "directionsText": body if not substantive else None,
                "passageText": body if substantive else None,
                "contextRequired": substantive,
                "stimulusSourcePage": source_page,
                "requiresSourceVisual": visual,
            }
    return assignments


def extract_section_questions(
    section_text: str,
    section: Section,
    pdf_name: str,
    answers: dict[int, str],
    default_page: int,
) -> dict[int, dict[str, Any]]:
    starts: dict[int, int] = {}
    content_starts: dict[int, int] = {}
    cursor = 0
    for question_number in range(1, section.count + 1):
        marker = question_marker(section_text, question_number, cursor)
        starts[question_number] = marker.start()
        content_starts[question_number] = marker.end()
        cursor = marker.end()

    contexts = parse_direction_ranges(
        section_text, starts, pdf_name, section, default_page
    )
    questions: dict[int, dict[str, Any]] = {}
    for question_number in range(1, section.count + 1):
        chunk_end = (
            starts[question_number + 1]
            if question_number < section.count
            else len(section_text)
        )
        raw_chunk = section_text[content_starts[question_number] : chunk_end]
        next_directions = re.search(
            r"(?im)^\s*Directions?\s+for\s+questions?", raw_chunk
        )
        if next_directions:
            raw_chunk = raw_chunk[: next_directions.start()]

        source_page = page_before(section_text, starts[question_number], default_page)
        answer_raw = compact(
            ANSWER_OVERRIDES.get(
                (pdf_name, section.name, question_number),
                answers.get(question_number, ""),
            )
        )
        answer_raw = re.sub(r"^([A-E])(?:\s*\[)+\s*$", r"\1", answer_raw)
        is_multi_part = bool(
            re.fullmatch(r"\(i\)\s*.+?\s*\(ii\)\s*.+", answer_raw, re.I)
        )
        sub_questions: list[dict[str, Any]] = []
        if is_multi_part:
            question_text, sub_questions = split_multi_part_question(
                raw_chunk, answer_raw
            )
            options = []
        else:
            question_text, options = split_options(raw_chunk)
        if not is_multi_part and len(options) < 2:
            numbered_question, numbered_options = split_numbered_options(raw_chunk)
            if len(numbered_options) >= 2:
                question_text, options = numbered_question, numbered_options

        context = contexts.get(
            question_number,
            {
                "stimulusId": None,
                "stimulusType": None,
                "directionsText": None,
                "passageText": None,
                "contextRequired": False,
                "stimulusSourcePage": None,
                "requiresSourceVisual": False,
            },
        )
        if not is_multi_part and len(options) < 2 and context.get("passageText"):
            _, fixed_options = split_options(context["passageText"])
            if len(fixed_options) >= 2:
                options = fixed_options

        if not question_text and context.get("directionsText"):
            question_text = context["directionsText"]

        choice_match = re.fullmatch(r"([A-E])", answer_raw)
        question_type = (
            "MULTI_PART"
            if is_multi_part
            else ("MCQ" if len(options) >= 2 else "NUMERIC")
        )
        if question_type == "MCQ" and not choice_match:
            raise ValueError(
                f"{pdf_name} {section.name} Q{question_number}: "
                f"choices parsed but answer is {answer_raw!r}"
            )
        correct_option = choice_match.group(1) if question_type == "MCQ" else None
        numeric_answer = answer_raw if question_type == "NUMERIC" else None
        visual_text = f"{question_text} {context.get('passageText') or ''}".lower()
        requires_visual = context["requiresSourceVisual"] or any(
            cue in visual_text for cue in VISUAL_CUES
        )
        if not question_text:
            question_text = (
                "Refer to the source figure and select the correct answer."
                if requires_visual
                else "Select the correct statement using the shared information."
            )

        questions[question_number] = {
            "sourceSection": section.name,
            "sourceQuestionNo": question_number,
            "sourcePage": source_page,
            "questionText": question_text,
            "options": options,
            "subQuestions": sub_questions,
            "questionType": question_type,
            "correctOption": correct_option,
            "numericAnswer": numeric_answer,
            "answerKeyRaw": answer_raw,
            **context,
            "requiresSourceVisual": requires_visual,
        }

    # Restore dependencies expressed as "the above problem/question".
    for question_number, item in questions.items():
        lowered = item["questionText"].lower()
        if question_number <= 1:
            continue
        if re.search(r"\b(?:in|from|using)\s+the\s+above\s+(?:problem|question)", lowered):
            previous = questions[question_number - 1]
            item["stimulusId"] = (
                f"{slug(pdf_name)}-{slug(section.name)}-linked-q{question_number - 1}"
            )
            item["stimulusType"] = "linked-question"
            item["passageText"] = (
                f"Referenced source problem (Question {question_number - 1}): "
                f"{previous['questionText']}"
            )
            item["contextRequired"] = True
            item["stimulusSourcePage"] = previous["sourcePage"]
            item["requiresSourceVisual"] = (
                item["requiresSourceVisual"] or previous["requiresSourceVisual"]
            )

    return questions


def parse_location(location: str) -> tuple[str, int]:
    match = re.fullmatch(r"(.+?)\s+Q(\d+)", location.strip())
    if not match:
        raise ValueError(f"Unrecognized schedule location: {location}")
    return match.group(1).strip(), int(match.group(2))


def load_schedule() -> list[dict[str, Any]]:
    workbook = load_workbook(SCHEDULE_PATH, read_only=True, data_only=True)
    sheet = workbook["Master Questions List"]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        index for index, row in enumerate(rows) if row and row[0] == "S.No"
    )
    schedule: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not row or row[0] is None:
            continue
        day_number = int(re.search(r"\d+", str(row[1])).group())
        daily_number = int(re.search(r"\d+", str(row[2])).group())
        schedule.append(
            {
                "id": int(row[0]),
                "day": day_number,
                "dayStr": str(row[1]),
                "dailyQNo": str(row[2]),
                "dailyNumber": daily_number,
                "topic": str(row[3]),
                "location": str(row[4]),
                "pdfFile": str(row[5]),
            }
        )
    return schedule


def difficulty(day: int, daily_number: int) -> tuple[int, str]:
    if daily_number <= 12 and day <= 10:
        return 1, "Foundational"
    if daily_number > 28 or day >= 22:
        return 3, "Advanced Benchmark"
    return 2, "Exam Standard"


def main() -> None:
    source_questions: dict[tuple[str, str, int], dict[str, Any]] = {}
    source_counts: dict[str, int] = {}

    for pdf_name, sections in MANIFEST.items():
        pdf_path = QUANT_DIR / pdf_name
        full_text, page_count = load_pdf_text(pdf_path)
        key_position = find_key_position(full_text)
        located_sections = locate_sections(full_text, sections, key_position)
        key_text = full_text[key_position:]

        for section_index, (section, start, end) in enumerate(located_sections):
            answers = parse_answer_section(key_text, section, section_index, sections)
            if len(answers) != section.count:
                missing = sorted(set(range(1, section.count + 1)) - set(answers))
                raise ValueError(
                    f"{pdf_name} {section.name}: missing answer keys {missing}"
                )
            parsed = extract_section_questions(
                full_text[start:end],
                section,
                pdf_name,
                answers,
                page_before(full_text, start),
            )
            for question_number, item in parsed.items():
                source_questions[(pdf_name, section.name, question_number)] = item

        source_counts[pdf_name] = page_count

    schedule = load_schedule()
    output: list[dict[str, Any]] = []
    unresolved: list[str] = []
    for scheduled in schedule:
        section_name, source_question_number = parse_location(scheduled["location"])
        key = (scheduled["pdfFile"], section_name, source_question_number)
        source = source_questions.get(key)
        if source is None:
            unresolved.append(
                f"{scheduled['id']}: {scheduled['pdfFile']} / {scheduled['location']}"
            )
            continue

        source = dict(source)
        if scheduled["topic"] in PURE_QUANT_TOPICS:
            source = normalize_quant_content(source)
            source.update(CONTENT_OVERRIDES.get(key, {}))

        level, label = difficulty(scheduled["day"], scheduled["dailyNumber"])
        source_pdf_url = f"/quant-sources/{scheduled['pdfFile']}"
        if source["questionType"] == "MCQ":
            answer_label = f"Choice {source['correctOption']}"
        elif source["questionType"] == "MULTI_PART":
            answer_label = source["answerKeyRaw"]
        else:
            answer_label = source["numericAnswer"]
        output.append(
            {
                "id": scheduled["id"],
                "day": scheduled["day"],
                "dayStr": scheduled["dayStr"],
                "dailyQNo": scheduled["dailyQNo"],
                "topic": scheduled["topic"],
                "category": CATEGORY_BY_TOPIC.get(
                    scheduled["topic"], "Quant & Logical Reasoning"
                ),
                "location": scheduled["location"],
                "pdfFile": scheduled["pdfFile"],
                "sourcePdfUrl": source_pdf_url,
                "sourceSection": source["sourceSection"],
                "sourceQuestionNo": source["sourceQuestionNo"],
                "sourcePage": source["sourcePage"],
                "difficultyLevel": level,
                "difficultyLabel": label,
                "questionType": source["questionType"],
                "questionText": source["questionText"],
                "directionsText": source["directionsText"],
                "passageText": source["passageText"],
                "stimulusId": source["stimulusId"],
                "stimulusType": source["stimulusType"],
                "stimulusSourcePage": source["stimulusSourcePage"],
                "contextRequired": source["contextRequired"],
                "requiresSourceVisual": source["requiresSourceVisual"],
                "imageUrl": None,
                "options": source["options"],
                "subQuestions": source["subQuestions"],
                "correctOption": source["correctOption"],
                "numericAnswer": source["numericAnswer"],
                "answerKeyRaw": source["answerKeyRaw"],
                "solution": f"Official source answer key: {answer_label}.",
                "whereThingsWentWrong": (
                    "Re-check every condition in the source question and any shared "
                    "passage, rule table or diagram before selecting an answer."
                ),
                "conceptTip": (
                    f"Source: {scheduled['location']} in {scheduled['pdfFile']}."
                ),
            }
        )

    if unresolved:
        raise ValueError("Unresolved schedule rows:\n" + "\n".join(unresolved))

    if len(output) != 1230:
        raise ValueError(f"Expected 1230 scheduled questions, generated {len(output)}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    PUBLIC_PDF_DIR.mkdir(parents=True, exist_ok=True)
    for pdf_name in MANIFEST:
        shutil.copy2(QUANT_DIR / pdf_name, PUBLIC_PDF_DIR / pdf_name)

    context_questions = sum(bool(item["contextRequired"]) for item in output)
    context_groups = len(
        {
            item["stimulusId"]
            for item in output
            if item["contextRequired"] and item["stimulusId"]
        }
    )
    numeric_questions = sum(item["questionType"] == "NUMERIC" for item in output)
    visual_questions = sum(bool(item["requiresSourceVisual"]) for item in output)
    print(
        json.dumps(
            {
                "questions": len(output),
                "sourcePdfs": len(source_counts),
                "contextGroups": context_groups,
                "questionsWithContext": context_questions,
                "numericQuestions": numeric_questions,
                "visualQuestions": visual_questions,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
